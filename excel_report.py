# ---------------------------------------------------------
# excel_report.py
# Excel reporting for aws-audit-tool
# ---------------------------------------------------------

from __future__ import annotations

# Imports -----------------------------------------------

import re
from datetime import datetime
from typing import Any, Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Constants ---------------------------------------------

EXCEL_TABLE_STYLE = "TableStyleMedium18"

# Data Sanitization -------------------------------------

# Excel doesn't handle timezone-aware datetimes, so we need to strip tzinfo before writing to Excel
def sanitize_for_excel(obj: Any) -> Any:
    if isinstance(obj, datetime) and obj.tzinfo is not None:
        return obj.replace(tzinfo=None)
    if isinstance(obj, dict):
        return {k: sanitize_for_excel(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [sanitize_for_excel(x) for x in obj]
    return obj


# Internal Helpers --------------------------------------

# Convert list of AWS tags (dicts with "Key" and "Value") into a simple dict for easier access
def _tags_dict(tags_list) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for t in tags_list or []:
        k = t.get("Key")
        v = t.get("Value")
        if k:
            out[k] = v or ""
    return out

# Excel table names must start with a letter or underscore, and can only contain letters, numbers, and underscores. 
# This function converts arbitrary sheet names into safe table names.
def _safe_table_name(name: str) -> str:
    n = re.sub(r"[^A-Za-z0-9_]", "_", name or "Sheet")
    if not re.match(r"^[A-Za-z_]", n):
        n = "_" + n
    return n[:250]

# Worksheet Formatting ----------------------------------

# Format the given worksheet as an Excel table, and auto-size columns. This makes the data look nicer and more readable in Excel.
def format_sheet_as_table(ws, table_style: str = EXCEL_TABLE_STYLE) -> None:
    max_row = ws.max_row
    max_col = ws.max_column

    if max_row >= 1:
        ws.freeze_panes = "A2"

    if max_row < 2 or max_col < 1:
        return

    # Excel tables require a reference like "A1:D10". 
    # We calculate the last column letter based on the max column number, and construct the table reference accordingly
    last_col_letter = get_column_letter(max_col)
    table_ref = f"A1:{last_col_letter}{max_row}"
    table_name = _safe_table_name(f"{ws.title}_Table")

    existing = {t.displayName for t in ws._tables}
    if table_name in existing:
        table_name = f"{table_name}_1"

    # Create the table with the specified style and add it to the worksheet. 
    # The style controls the colors and formatting of the table in Excel
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name=table_style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    # Auto-size columns based on the maximum length of the content in each column. We iterate through each column
    # Check the length of the content in each cell, and set the column width accordingly (with a max width to prevent excessively wide columns)
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

# Apply formatting to all worksheets in the workbook. 
# This is called after writing all the data to Excel, to ensure that each sheet is formatted as a table and has auto-sized columns.
def apply_workbook_formatting(xlsx_path: str) -> None:
    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        format_sheet_as_table(ws, table_style=EXCEL_TABLE_STYLE)
    wb.save(xlsx_path)

# Report Builder ----------------------------------------

# The main function to build the Excel report. It takes the inventory data (as a dictionary) and the output path for the Excel file. 
# It processes the inventory data, extracts relevant information for each AWS resource type, and organizes it into separate lists of dictionaries (one list per resource type). 
# Each dictionary in the list represents a row in the Excel sheet, with keys corresponding to column names.
def build_excel(inventory: Dict[str, Any], out_xlsx_path: str) -> None:
    items = inventory.get("items", []) or []
    global_block = inventory.get("global", {}) or {}
    run_info = inventory.get("run_info", {}) or {}

    s3_buckets_all = global_block.get("s3_buckets", []) or []
    iam_users_all = global_block.get("iam_users", []) or []
    account_aliases = global_block.get("account_aliases", []) or []

    vpcs_all, subnets_all, rts_all, sgs_all, ec2_all, lbs_all = [], [], [], [], [], []
    ebs_vols_all, rds_instances_all, rds_clusters_all, asgs_all, dynamodb_all = [], [], [], [], []

    for region_block in items:
        region = region_block.get("region")

        # VPCs, Subnets, Route Tables, Security Groups, EC2 Instances, Load Balancers, EBS Volumes, RDS Instances, RDS Clusters, AutoScaling Groups, DynamoDB Tables, etc. 
        # We extract relevant fields for each resource type and append them to the corresponding list.
        for v in region_block.get("vpcs", []):
            vpcs_all.append({
                "Region": region,
                "VpcId": v.get("VpcId"),
                "CidrBlock": v.get("CidrBlock"),
                "IsDefault": v.get("IsDefault"),
                "State": v.get("State"),
            })

        for s in region_block.get("subnets", []):
            subnets_all.append({
                "Region": region,
                "SubnetId": s.get("SubnetId"),
                "VpcId": s.get("VpcId"),
                "CidrBlock": s.get("CidrBlock"),
                "AvailabilityZone": s.get("AvailabilityZone"),
                "State": s.get("State"),
            })

        for rt in region_block.get("route_tables", []):
            rts_all.append({
                "Region": region,
                "RouteTableId": rt.get("RouteTableId"),
                "VpcId": rt.get("VpcId"),
                "Associations": len(rt.get("Associations", []) or []),
                "Routes": len(rt.get("Routes", []) or []),
            })

        for sg in region_block.get("security_groups", []):
            sgs_all.append({
                "Region": region,
                "GroupId": sg.get("GroupId"),
                "GroupName": sg.get("GroupName"),
                "VpcId": sg.get("VpcId"),
                "IngressRules": len(sg.get("IpPermissions", []) or []),
                "EgressRules": len(sg.get("IpPermissionsEgress", []) or []),
                "Description": sg.get("Description"),
            })

        for ins in region_block.get("instances", []):
            tags = _tags_dict(ins.get("Tags", []) or [])
            name = tags.get("Name", "")

            ec2_all.append({
                "Region": region,
                "InstanceId": ins.get("InstanceId"),
                "Name": name,
                "InstanceType": ins.get("InstanceType"),
                "State": (ins.get("State") or {}).get("Name"),
                "VpcId": ins.get("VpcId"),
                "SubnetId": ins.get("SubnetId"),
                "PrivateIp": ins.get("PrivateIpAddress"),
                "PublicIp": ins.get("PublicIpAddress"),
            })

        for lb in region_block.get("load_balancers", []):
            lbs_all.append({
                "Region": region,
                "LoadBalancerArn": lb.get("LoadBalancerArn"),
                "Name": lb.get("LoadBalancerName"),
                "Type": lb.get("Type"),
                "Scheme": lb.get("Scheme"),
                "VpcId": lb.get("VpcId"),
                "State": (lb.get("State") or {}).get("Code"),
                "DNSName": lb.get("DNSName"),
            })

        for v in region_block.get("ebs_volumes", []):
            atts = v.get("Attachments", []) or []
            instance_id = atts[0].get("InstanceId") if atts else ""
            device = atts[0].get("Device") if atts else ""

            ebs_vols_all.append({
                "Region": region,
                "VolumeId": v.get("VolumeId"),
                "Type": v.get("VolumeType"),
                "SizeGiB": v.get("Size"),
                "State": v.get("State"),
                "Encrypted": v.get("Encrypted"),
                "Iops": v.get("Iops"),
                "Throughput": v.get("Throughput"),
                "AvailabilityZone": v.get("AvailabilityZone"),
                "AttachedInstanceId": instance_id,
                "Device": device,
            })

        for db in region_block.get("rds_db_instances", []):
            rds_instances_all.append({
                "Region": region,
                "DBInstanceIdentifier": db.get("DBInstanceIdentifier"),
                "Engine": db.get("Engine"),
                "EngineVersion": db.get("EngineVersion"),
                "DBInstanceClass": db.get("DBInstanceClass"),
                "Status": db.get("DBInstanceStatus"),
                "MultiAZ": db.get("MultiAZ"),
                "StorageEncrypted": db.get("StorageEncrypted"),
                "PubliclyAccessible": db.get("PubliclyAccessible"),
                "BackupRetentionPeriod": db.get("BackupRetentionPeriod"),
                "AllocatedStorage": db.get("AllocatedStorage"),
                "Endpoint": (db.get("Endpoint") or {}).get("Address"),
            })

        for c in region_block.get("rds_db_clusters", []):
            rds_clusters_all.append({
                "Region": region,
                "DBClusterIdentifier": c.get("DBClusterIdentifier"),
                "Engine": c.get("Engine"),
                "EngineVersion": c.get("EngineVersion"),
                "Status": c.get("Status"),
                "MultiAZ": c.get("MultiAZ"),
                "Endpoint": c.get("Endpoint"),
                "ReaderEndpoint": c.get("ReaderEndpoint"),
            })

        for a in region_block.get("autoscaling_groups", []):
            asgs_all.append({
                "Region": region,
                "AutoScalingGroupName": a.get("AutoScalingGroupName"),
                "MinSize": a.get("MinSize"),
                "MaxSize": a.get("MaxSize"),
                "DesiredCapacity": a.get("DesiredCapacity"),
                "VpcZoneIdentifier": a.get("VPCZoneIdentifier"),
                "LaunchTemplate": (a.get("LaunchTemplate") or {}).get("LaunchTemplateName"),
                "Instances": len(a.get("Instances", []) or []),
            })

        for t in region_block.get("dynamodb_tables", []):
            dynamodb_all.append({
                "Region": region,
                "TableName": t.get("TableName"),
                "Status": t.get("TableStatus"),
                "BillingMode": (t.get("BillingModeSummary") or {}).get("BillingMode"),
                "ItemCount": t.get("ItemCount"),
                "SizeBytes": t.get("TableSizeBytes"),
                "Arn": t.get("TableArn"),
            })

    run_info_rows = [{
        "timestamp_utc": run_info.get("timestamp_utc", ""),
        "regions": ", ".join(run_info.get("regions", []) or []),
        "account_name": run_info.get("account_name", ""),
        "account_id_env": run_info.get("account_id_env", ""),
        "sts_account": (run_info.get("sts_identity", {}) or {}).get("Account", ""),
        "sts_arn": (run_info.get("sts_identity", {}) or {}).get("Arn", ""),
        "account_aliases": ", ".join(account_aliases) if account_aliases else "",
    }]

    summary = [{
        "VPCs": len(vpcs_all),
        "Subnets": len(subnets_all),
        "RouteTables": len(rts_all),
        "SecurityGroups": len(sgs_all),
        "EC2Instances": len(ec2_all),
        "LoadBalancers": len(lbs_all),
        "EBSVolumes": len(ebs_vols_all),
        "RDSInstances": len(rds_instances_all),
        "RDSClusters": len(rds_clusters_all),
        "AutoScalingGroups": len(asgs_all),
        "DynamoDBTables": len(dynamodb_all),
        "S3Buckets": len(s3_buckets_all),
        "IAMUsers": len(iam_users_all),
    }]

    # Convert the lists of dictionaries into pandas DataFrames, sanitize the data for Excel, and write each DataFrame to a separate sheet in the Excel file
    def df(records: Any) -> pd.DataFrame:
        return pd.DataFrame(sanitize_for_excel(records))

    with pd.ExcelWriter(out_xlsx_path, engine="openpyxl") as writer:
        df(vpcs_all).to_excel(writer, index=False, sheet_name="VPCs")
        df(subnets_all).to_excel(writer, index=False, sheet_name="Subnets")
        df(rts_all).to_excel(writer, index=False, sheet_name="RouteTables")
        df(sgs_all).to_excel(writer, index=False, sheet_name="SecurityGroups")
        df(ec2_all).to_excel(writer, index=False, sheet_name="EC2")
        df(lbs_all).to_excel(writer, index=False, sheet_name="LoadBalancers")

        df(ebs_vols_all).to_excel(writer, index=False, sheet_name="EBS_Volumes")
        df(rds_instances_all).to_excel(writer, index=False, sheet_name="RDS_Instances")
        df(rds_clusters_all).to_excel(writer, index=False, sheet_name="RDS_Clusters")
        df(asgs_all).to_excel(writer, index=False, sheet_name="AutoScaling")
        df(dynamodb_all).to_excel(writer, index=False, sheet_name="DynamoDB")

        df(s3_buckets_all).to_excel(writer, index=False, sheet_name="S3_Buckets")
        df(iam_users_all).to_excel(writer, index=False, sheet_name="IAM_Users")

        df(run_info_rows).to_excel(writer, index=False, sheet_name="Run_Info")
        df(summary).to_excel(writer, index=False, sheet_name="Summary")

    apply_workbook_formatting(out_xlsx_path)
